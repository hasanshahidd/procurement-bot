import os
from openpyxl import load_workbook

def safe_str(value):
    if value is None:
        return ""
    return str(value).strip()

def safe_float(value):
    if value is None:
        return 0.0
    try:
        return float(value)
    except (ValueError, TypeError):
        return 0.0

def safe_int(value):
    if value is None:
        return None
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return None

def load_excel_data(file_path: str) -> list:
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")
    
    wb = load_workbook(file_path, data_only=True)
    sheet = wb.active
    
    headers = [cell.value for cell in sheet[1]]
    records = []
    
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
            
        record = {
            "year": safe_int(row[0]) or 2026,
            "quarter": safe_str(row[1]) or "Q1",
            "month": safe_str(row[2]) or "January",
            "period": safe_str(row[3]) or "2026-01",
            "date": safe_str(row[4]) or "2026-01-01",
            "pr_number": safe_str(row[5]) or f"PR-{row_idx}",
            "description": safe_str(row[6]) or "No description",
            "department": safe_str(row[7]) or "Unknown",
            "contact_person": safe_str(row[8]) or "Unknown",
            "assign_to": safe_str(row[9]) or "Unassigned",
            "budget": safe_float(row[10]),
            "budget_q1": safe_float(row[11]),
            "budget_q2": safe_float(row[12]),
            "budget_q3": safe_float(row[13]),
            "budget_q4": safe_float(row[14]),
            "source_method": safe_str(row[15]) or "Unknown",
            "status": safe_str(row[16]) or "Pending",
            "supplier_details": safe_str(row[17]) if len(row) > 17 else None,
            "supplier_rating": safe_str(row[18]) if len(row) > 18 else None,
            "local_content_percentage": safe_float(row[19]) if len(row) > 19 else None,
            "pr_approval_scope_input": safe_str(row[20]) if len(row) > 20 else None,
            "approving_authority": safe_str(row[21]) if len(row) > 21 else None,
            "planned": safe_str(row[22]) if len(row) > 22 else None,
            "target_date": safe_str(row[23]) if len(row) > 23 else None,
            "actual_project_start": safe_str(row[24]) if len(row) > 24 else None,
            "sla": safe_int(row[25]) if len(row) > 25 else None,
            "note": safe_str(row[26]) if len(row) > 26 else None,
            "review_approval_scope_eval": safe_float(row[27]) if len(row) > 27 else None,
            "floating": safe_float(row[28]) if len(row) > 28 else None,
            "tender_submit_by_vendor": safe_float(row[29]) if len(row) > 29 else None,
            "evaluation": safe_float(row[30]) if len(row) > 30 else None,
            "award_approval": safe_float(row[31]) if len(row) > 31 else None,
            "contract_and_po": safe_float(row[32]) if len(row) > 32 else None,
            "pr_approval_scope_input_pd": safe_float(row[33]) if len(row) > 33 else None,
            "review_approval_scope_eval_pd": safe_float(row[34]) if len(row) > 34 else None,
            "floating_pd": safe_float(row[35]) if len(row) > 35 else None,
            "tender_submit_by_vendor_pd": safe_float(row[36]) if len(row) > 36 else None,
            "evaluation_pd": safe_float(row[37]) if len(row) > 37 else None,
            "award_approval_pd": safe_float(row[38]) if len(row) > 38 else None,
            "contract_and_po_pd": safe_float(row[39]) if len(row) > 39 else None,
            "total_days_pd": safe_float(row[40]) if len(row) > 40 else None,
            "review_approval_scope_eval_ad": safe_float(row[41]) if len(row) > 41 else None,
            "floating_ad": safe_float(row[42]) if len(row) > 42 else None,
            "tender_submit_by_vendor_ad": safe_float(row[43]) if len(row) > 43 else None,
            "evaluation_ad": safe_float(row[44]) if len(row) > 44 else None,
            "award_approval_ad": safe_float(row[45]) if len(row) > 45 else None,
            "contract_and_po_ad": safe_float(row[46]) if len(row) > 46 else None,
            "total_days_ad": safe_float(row[47]) if len(row) > 47 else None,
            "review_approval_scope_eval_pd_sla": safe_float(row[48]) if len(row) > 48 else None,
            "floating_pd_sla": safe_float(row[49]) if len(row) > 49 else None,
            "tender_submit_by_vendor_pd_sla": safe_float(row[50]) if len(row) > 50 else None,
            "evaluation_pd_sla": safe_float(row[51]) if len(row) > 51 else None,
            "award_approval_pd_sla": safe_float(row[52]) if len(row) > 52 else None,
            "contract_and_po_pd_sla": safe_float(row[53]) if len(row) > 53 else None,
            "review_approval_scope_eval_ad_sla": safe_float(row[54]) if len(row) > 54 else None,
            "floating_ad_sla": safe_float(row[55]) if len(row) > 55 else None,
            "tender_submit_by_vendor_ad_sla": safe_float(row[56]) if len(row) > 56 else None,
            "evaluation_ad_sla": safe_float(row[57]) if len(row) > 57 else None,
            "award_approval_ad_sla": safe_float(row[58]) if len(row) > 58 else None,
            "contract_and_po_ad_sla": safe_float(row[59]) if len(row) > 59 else None,
            "review_approval_scope_eval_diff_sla": safe_float(row[60]) if len(row) > 60 else None,
            "floating_diff_sla": safe_float(row[61]) if len(row) > 61 else None,
            "tender_submit_by_vendor_diff_sla": safe_float(row[62]) if len(row) > 62 else None,
            "evaluation_diff_sla": safe_float(row[63]) if len(row) > 63 else None,
            "award_approval_diff_sla": safe_float(row[64]) if len(row) > 64 else None,
            "contract_and_po_diff_sla": safe_float(row[65]) if len(row) > 65 else None,
            "project_status": safe_int(row[66]) if len(row) > 66 else None,
            "risk": safe_str(row[67]) if len(row) > 67 else None,  # TEXT: "Low", "Medium", "High"
            "duration": safe_int(row[68]) if len(row) > 68 else None,
            "last_status_date": safe_str(row[69]) if len(row) > 69 else None,
            "status_duration": safe_int(row[70]) if len(row) > 70 else None,
            "status_sla": safe_int(row[71]) if len(row) > 71 else None,
            "status_co": safe_str(row[72]) if len(row) > 72 else None,
            "escalate_48h": safe_str(row[73]) if len(row) > 73 else None,
            "ceo_escalation": safe_str(row[74]) if len(row) > 74 else None,
        }
            
        records.append(record)
    
    return records
